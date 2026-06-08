from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from datetime import datetime, timedelta
import logging
import os
import socket
import re

from database import init_db, get_conn
from imweb_api import get_paid_orders, extract_order_info, set_order_to_standby
from sms_parser import parse_sms

import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    stream=sys.stdout
)
logger = logging.getLogger(__name__)


# ── 한국시간(KST, UTC+9) 헬퍼: Railway 서버는 UTC라 표시/기록 시간이 9시간 어긋남 ──
def now_kst():
    """서버(UTC) 기준에서 한국시간 naive datetime 반환"""
    return datetime.utcnow() + timedelta(hours=9)

app = Flask(__name__, static_folder='static')
CORS(app)


# ══════════════════════════════════════════════════════════════════
#  ★ 요청 추적 미들웨어 (디버깅용)
#  모든 들어오는 요청을 무조건 기록 → Automate/외부에서 보낸 요청이
#  서버에 닿기만 하면 Railway 로그에 한 줄이 무조건 찍힘.
# ══════════════════════════════════════════════════════════════════
@app.before_request
def _log_every_request():
    try:
        # 정적 파일이나 favicon 등 시끄러운 요청은 제외
        if request.path in ('/favicon.ico',):
            return
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        ctype = request.content_type or ''
        clen = request.content_length or 0
        # 짧은 본문이면 내용도 같이 (SMS POST 디버깅 핵심)
        body_preview = ''
        if request.method == 'POST' and clen < 2000:
            try:
                raw = request.get_data(cache=True, as_text=True) or ''
                body_preview = ' body=' + raw.replace('\n', '\\n')[:300]
            except Exception:
                body_preview = ' body=<read-failed>'
        logger.info(
            f"➡️  {request.method} {request.path} from={client_ip} "
            f"ct={ctype} len={clen}{body_preview}"
        )
    except Exception as e:
        logger.error(f"요청 로깅 오류: {e}")


# ══════════════════════════════════════════════════════════════════
#  헬스체크 (외부 모니터링/연결 테스트용)
# ══════════════════════════════════════════════════════════════════
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'ok': True, 'service': 'jiyang-haworthia', 'time': now_kst().isoformat()})


# ══════════════════════════════════════════════════════════════════
#  메인 페이지 서빙
# ══════════════════════════════════════════════════════════════════
@app.route('/')
def index():
    response = send_file('index.html')
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# ══════════════════════════════════════════════════════════════════
#  SMS Webhook  ← Automate/MacroDroid가 여기로 POST 전송
#  여러 경로 모두 같은 핸들러로: /sms, /api/sms, /webhook/sms
# ══════════════════════════════════════════════════════════════════
@app.route('/sms',         methods=['POST'])
@app.route('/api/sms',     methods=['POST'])
@app.route('/webhook/sms', methods=['POST'])
def receive_sms():
    try:
        # JSON 우선, 실패 시 form/text도 시도 (Automate 설정 차이 대응)
        data = request.get_json(silent=True)
        if not data:
            # JSON 헤더 없이 form-encoded로 보내는 경우
            if request.form:
                data = request.form.to_dict()
            else:
                # raw text를 body로 취급
                raw = request.get_data(as_text=True) or ''
                if raw.strip():
                    data = {'body': raw}
                else:
                    data = {}

        body      = (data.get('body')   or data.get('message') or data.get('text')   or '').strip()
        sender    = (data.get('sender') or data.get('from')    or data.get('source') or '').strip()
        recv_time = data.get('time') or now_kst().isoformat()

        if not body:
            logger.warning(f"⚠️  SMS 본문 없음 (data keys: {list(data.keys()) if data else 'none'})")
            return jsonify({'ok': False, 'error': 'body 필드가 비어있음'}), 400

        logger.info(f"📨 SMS 수신: sender='{sender}' body='{body[:80]}'")

        parsed = parse_sms(body)
        sms_id = None
        try:
            conn = get_conn()
            cur = conn.execute(
                '''INSERT INTO sms_payments (sender, body, parsed_name, parsed_amount, received_at)
                   VALUES (?, ?, ?, ?, ?)''',
                (sender, body,
                 parsed['name']   if parsed else None,
                 parsed['amount'] if parsed else None,
                 recv_time)
            )
            sms_id = cur.lastrowid
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"SMS DB 저장 오류: {e}")

        if parsed:
            try:
                match_sms_to_order(parsed, recv_time, sms_id=sms_id)
                logger.info(f"✅ SMS 파싱 성공: {parsed['bank']} {parsed.get('name','?')} {parsed['amount']:,}원")
            except Exception as e:
                logger.error(f"SMS 매칭 오류: {e}")
        else:
            logger.info(f"ℹ️  SMS 파싱 불가: {body[:60]}")

        return jsonify({'ok': True, 'received': True, 'parsed': bool(parsed), 'sms_id': sms_id})

    except Exception as e:
        logger.exception(f"SMS 수신 처리 예외: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 200


# ══════════════════════════════════════════════════════════════════
#  거래명세표 업로드
# ══════════════════════════════════════════════════════════════════
def extract_live_date_from_filename(filename):
    """파일명에서 라이브방송 날짜(YYYY-MM-DD) 추출"""
    if not filename: return None
    m = re.search(r'(20\d{2})-(\d{2})-(\d{2})', filename)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r'(20\d{2})(\d{2})(\d{2})', filename)
    if m:
        try:
            datetime.strptime(f"{m.group(1)}-{m.group(2)}-{m.group(3)}", '%Y-%m-%d')
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
        except ValueError:
            return None
    return None


@app.route('/api/session/upload', methods=['POST'])
def upload_session():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']

    # live_date 결정: 1) 폼 입력값  2) 파일명 추출  3) 오늘 (최후)
    live_date_form = (request.form.get('live_date') or '').strip()
    if live_date_form:
        live_date = live_date_form
        live_date_source = 'form'
    else:
        extracted = extract_live_date_from_filename(file.filename)
        if extracted:
            live_date = extracted
            live_date_source = 'filename'
        else:
            live_date = now_kst().strftime('%Y-%m-%d')
            live_date_source = 'today_fallback'

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        orders = parse_invoice_excel(ws)
        # 🌿 판매품 리스트 시트도 자동 파싱
        catalog_items = []
        for sn in wb.sheetnames:
            if '판매품' in sn or '재고' in sn:
                try:
                    catalog_items = parse_product_catalog(wb[sn])
                    logger.info(f"  📚 판매품 리스트 시트 발견: '{sn}' ({len(catalog_items)}건)")
                    break
                except Exception as ce:
                    logger.warning(f"판매품 시트 파싱 실패: {ce}")
    except Exception as e:
        return jsonify({'error': f'엑셀 파싱 오류: {e}'}), 400

    try:
        live_dt = datetime.strptime(live_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': f'라이브 날짜 형식 오류: {live_date}'}), 400

    # 7일 자동확인: 라방날짜 +1일부터 시작
    check_start = (live_dt + timedelta(days=1)).strftime('%Y-%m-%d')
    check_end   = (live_dt + timedelta(days=8)).strftime('%Y-%m-%d')
    logger.info(f"📅 live_date 결정: {live_date} (source={live_date_source})")

    conn = get_conn()
    c = conn.execute(
        '''INSERT INTO live_sessions (filename, live_date, check_start, check_end, created_at)
           VALUES (?, ?, ?, ?, ?)''',
        (file.filename, live_date, check_start, check_end, now_kst().isoformat())
    )
    session_id = c.lastrowid

    for o in orders:
        conn.execute(
            '''INSERT INTO orders (session_id, buyer_name, item, item_no, amount, pay_type, status)
               VALUES (?, ?, ?, ?, ?, ?, 'pending')''',
            (session_id, o['name'], o.get('item', ''), o.get('item_no', ''), o['amount'], o.get('pay_type', ''))
        )

    # 판매품 리스트 백업 — 같은 세션 중복 방지
    try:
        existing_names = set(r[0] for r in conn.execute("SELECT DISTINCT item_name FROM product_catalog").fetchall())
        for p in catalog_items:
            nm = (p.get('item_name') or '').strip()
            if not _catalog_keep(p.get('price')):   # 50만원 이하/가격없음 제외
                continue
            if nm in existing_names:                 # 같은 이름 중복 제외
                continue
            existing_names.add(nm)
            conn.execute(
                '''INSERT INTO product_catalog (session_id, live_date, item_no, item_name, price, remaining)
                   VALUES (?, ?, ?, ?, ?, ?)''',
                (session_id, live_date, p.get('item_no',''), nm, p.get('price'), p.get('remaining',''))
            )
        if catalog_items:
            logger.info(f"  📚 판매품 카탈로그 저장: {len(catalog_items)}건 (session={session_id})")
    except Exception as e:
        logger.warning(f"판매품 카탈로그 저장 실패: {e}")

    conn.commit()
    conn.close()

    logger.info(f"📋 세션 등록: {file.filename} ({len(orders)}명, {check_start}~{check_end})")

    # 업로드 즉시 한 번 확인 실행
    run_auto_check()

    return jsonify({
        'ok': True,
        'session_id': session_id,
        'orders_count': len(orders),
        'check_start': check_start,
        'check_end': check_end
    })


# 판매품 카탈로그 저장 기준: 이 가격 이하(또는 가격없음)는 저장 안 함 (렉 방지)
CATALOG_MIN_PRICE = 500000
def _catalog_keep(price):
    try:
        return price is not None and int(price) > CATALOG_MIN_PRICE
    except Exception:
        return False


def parse_product_catalog(ws):
    """거래명세표의 '판매품 리스트' 시트에서 (번호, 이름, 가격, 잔여) 추출.
    헤더 자동 감지: 번호/연번/no, 하월시아/품명/상품/이름, 가격/금액, 잔여/재고"""
    rows = list(ws.values)
    result = []
    NO_HINTS    = ['번호', '연번', 'no']
    NAME_HINTS  = ['하월시아', '품명', '상품', '식물', '이름']
    PRICE_HINTS = ['가격', '금액', '판매']
    REM_HINTS   = ['잔여', '재고', '수량', '개수']
    header_idx = no_col = name_col = price_col = rem_col = -1
    for i, row in enumerate(rows[:15]):
        if not row: continue
        cells = [str(c or '').strip() for c in row]
        ni = next((j for j, c in enumerate(cells) if any(h in c.lower() if 'no' in h else h in c for h in NO_HINTS)), -1)
        nn = next((j for j, c in enumerate(cells) if any(h in c for h in NAME_HINTS)), -1)
        pp = next((j for j, c in enumerate(cells) if any(h in c for h in PRICE_HINTS)), -1)
        rr = next((j for j, c in enumerate(cells) if any(h in c for h in REM_HINTS)), -1)
        if ni >= 0 and nn >= 0 and pp >= 0:
            header_idx, no_col, name_col, price_col, rem_col = i, ni, nn, pp, rr
            break
    if header_idx < 0:
        return result
    for row in rows[header_idx+1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        row = list(row)
        no_v = row[no_col] if no_col < len(row) else None
        nm_v = row[name_col] if name_col < len(row) else None
        pr_v = row[price_col] if price_col < len(row) else None
        rm_v = row[rem_col] if rem_col >= 0 and rem_col < len(row) else None
        # 번호/이름이 둘 다 비어 있으면 스킵 (메모 텍스트 등)
        if (no_v is None or str(no_v).strip() == '') and (not nm_v or not str(nm_v).strip()):
            continue
        # 번호 정규화
        if isinstance(no_v, (int, float)) and float(no_v).is_integer():
            item_no = str(int(no_v))
        else:
            item_no = str(no_v).strip() if no_v is not None else ''
        item_name = str(nm_v or '').strip()
        # 가격
        try:
            raw = str(pr_v or '').replace(',', '').replace('원', '').strip()
            price = int(float(raw)) if raw else None
        except Exception:
            price = None
        remaining = str(rm_v).strip() if rm_v is not None and str(rm_v).strip() else ''
        result.append({'item_no': item_no, 'item_name': item_name, 'price': price, 'remaining': remaining})
    return result


def parse_invoice_excel(ws):
    """거래명세표 엑셀에서 구매자/번호/이름/금액 파싱.
    헤더 인식 키워드:
      구매자: '구매자', '닉네임', '성함', '주문자'
      번호:   '번호', '연번', 'No.', 'no.'
      이름:   '하월시아', '품명', '상품', '품목', '내역', '식물', '제품'
      금액:   '판매가격', '금액', '합계', '가격', '총액', '결제'
    """
    rows = list(ws.values)
    orders = []

    header_idx = name_col = amount_col = item_col = itemno_col = -1

    # '이름' 단독은 '하월시아 이름'에도 매치되므로 NAME_HINTS에 두지 않음
    NAME_HINTS   = ['구매자', '닉네임', '성함', '주문자']
    NUMBER_HINTS = ['번호', '연번', 'No.', 'no.']
    ITEM_HINTS   = ['하월시아', '품명', '상품', '품목', '내역', '식물', '제품']
    AMOUNT_HINTS = ['판매가격', '금액', '합계', '가격', '총액', '결제']

    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        cells = [str(c or '').strip() for c in row]
        ni = next((j for j, c in enumerate(cells) if any(h in c for h in NAME_HINTS)), -1)
        ai = next((j for j, c in enumerate(cells) if any(h in c for h in AMOUNT_HINTS)), -1)
        if ni >= 0 and ai >= 0:
            header_idx = i
            name_col   = ni
            amount_col = ai
            item_col   = next((j for j, c in enumerate(cells) if any(h in c for h in ITEM_HINTS)), -1)
            itemno_col = next((j for j, c in enumerate(cells) if any(h in c for h in NUMBER_HINTS)), -1)
            break

    if header_idx < 0:
        return orders

    for row in rows[header_idx + 1:]:
        if not row or all((c is None or str(c).strip() == '') for c in row):
            continue
        row = list(row)

        name = str(row[name_col] or '').strip() if name_col < len(row) else ''
        # 금액: 빈칸/비숫자는 0으로 처리 (0원·무료·선물 상품도 명세서에 포함)
        raw = str(row[amount_col] or '').replace(',', '').replace('원', '').strip() if amount_col < len(row) else ''
        try:
            amt = int(float(raw)) if raw != '' else 0
        except Exception:
            amt = 0
        item = str(row[item_col] or '').strip() if item_col >= 0 and item_col < len(row) else ''
        itemno = ''
        if itemno_col >= 0 and itemno_col < len(row):
            v = row[itemno_col]
            if v is not None and str(v).strip():
                try:
                    if isinstance(v, (int, float)) and float(v).is_integer():
                        itemno = str(int(v))
                    else:
                        itemno = str(v).strip()
                except Exception:
                    itemno = str(v).strip()

        # 총합계/소계/합계 같은 요약행 스킵
        clean = name.replace(' ', '').replace('\u00a0', '')
        if clean in {'총합계', '총계', '소계', '합계', '총합', '계', '총주문', '주문합계'}:
            continue
        # 구매자가 아닌 메모/상태 행 제외 (예: '문자발송', '입금 x')
        if any(k in clean for k in ('문자발송', '입금x', '입금X', '문자전송')):
            continue
        if name and (amt > 0 or item):
            orders.append({'name': name, 'item': item, 'item_no': itemno, 'amount': amt})

    return orders


# ══════════════════════════════════════════════════════════════════
#  세션 목록
# ══════════════════════════════════════════════════════════════════
@app.route('/api/sessions', methods=['GET'])
def get_sessions():
    conn = get_conn()
    sessions = conn.execute('SELECT * FROM live_sessions ORDER BY created_at DESC').fetchall()
    result = []
    for s in sessions:
        s = dict(s)
        orders = conn.execute('SELECT * FROM orders WHERE session_id=?', (s['id'],)).fetchall()
        s['orders']    = [dict(o) for o in orders]
        s['confirmed'] = sum(1 for o in s['orders'] if o['status'] == 'confirmed')
        s['pending']   = sum(1 for o in s['orders'] if o['status'] == 'pending')
        s['total']     = len(s['orders'])
        result.append(s)
    conn.close()
    return jsonify(result)


@app.route('/api/sessions/<int:session_id>', methods=['DELETE'])
def delete_session(session_id):
    """업로드한 거래명세표(세션)와 그에 딸린 모든 데이터를 삭제.
    orders / check_logs / match_candidates / buyer_status / product_catalog 연쇄 삭제."""
    conn = get_conn()
    try:
        sess = conn.execute('SELECT * FROM live_sessions WHERE id=?', (session_id,)).fetchone()
        if not sess:
            return jsonify({'error': '해당 세션이 없습니다'}), 404
        for tbl in ('orders', 'check_logs', 'match_candidates', 'buyer_status', 'product_catalog'):
            try:
                conn.execute(f'DELETE FROM {tbl} WHERE session_id=?', (session_id,))
            except Exception as te:
                logger.warning(f'{tbl} 삭제 중 경고: {te}')
        conn.execute('DELETE FROM live_sessions WHERE id=?', (session_id,))
        conn.commit()
        logger.info(f"🗑 세션 삭제: id={session_id} ({sess['filename']})")
        return jsonify({'ok': True, 'deleted_session': session_id})
    except Exception as e:
        logger.exception(f"세션 삭제 실패: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════
#  배송 목록
# ══════════════════════════════════════════════════════════════════
@app.route('/api/orders/pending', methods=['GET'])
def get_pending_orders():
    """미확인 대기 주문 - 구매자별 합산"""
    session_id = request.args.get('session_id')
    conn = get_conn()

    sql = '''SELECT o.buyer_name, 
                    GROUP_CONCAT(o.item, ' / ') as items,
                    SUM(o.amount) as total_amount,
                    COUNT(*) as item_count,
                    ls.live_date, ls.filename, o.session_id
             FROM orders o
             JOIN live_sessions ls ON o.session_id = ls.id
             WHERE o.status = 'pending'
               AND o.buyer_name NOT LIKE '%문자발송%'
               AND o.buyer_name NOT LIKE '%입금x%'
               AND o.buyer_name NOT LIKE '%입금 x%' '''
    params = []
    if session_id:
        sql += ' AND o.session_id = ?'
        params.append(session_id)
    sql += ' GROUP BY o.session_id, o.buyer_name ORDER BY o.buyer_name'

    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/orders/confirmed', methods=['GET'])
def get_confirmed_orders():
    """입금확인 완료 주문 - 구매자별 합산"""
    session_id = request.args.get('session_id')
    conn = get_conn()

    sql = '''SELECT o.buyer_name,
                    GROUP_CONCAT(o.item, ' / ') as items,
                    SUM(o.amount) as total_amount,
                    COUNT(*) as item_count,
                    MAX(o.pay_type) as pay_type,
                    MAX(o.confirmed_at) as confirmed_at,
                    ls.live_date, ls.filename, o.session_id
             FROM orders o
             JOIN live_sessions ls ON o.session_id = ls.id
             WHERE o.status = 'confirmed' '''
    params = []
    if session_id:
        sql += ' AND o.session_id = ?'
        params.append(session_id)
    sql += ' GROUP BY o.session_id, o.buyer_name ORDER BY o.confirmed_at DESC'

    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/orders/confirm-by-buyer', methods=['POST'])
def confirm_by_buyer():
    """구매자 이름으로 해당 세션의 모든 주문 수동 확인"""
    data = request.get_json(force=True) or {}
    session_id = data.get('session_id')
    buyer_name = data.get('buyer_name')
    if not session_id or not buyer_name:
        return jsonify({'error': '필수 파라미터 없음'}), 400
    conn = get_conn()
    conn.execute(
        "UPDATE orders SET status='confirmed', confirmed_at=?, pay_type='manual' WHERE session_id=? AND buyer_name=? AND status='pending'",
        (now_kst().isoformat(), session_id, buyer_name))
    conn.commit()
    conn.close()
    logger.info(f"수동확인: {buyer_name} (session {session_id})")
    return jsonify({'ok': True})



def manual_confirm_order(order_id):
    """수동으로 입금확인 처리"""
    conn = get_conn()
    conn.execute('''UPDATE orders SET status='confirmed', confirmed_at=?, pay_type='manual'
                   WHERE id=?''', (now_kst().isoformat(), order_id))
    conn.commit()
    conn.close()
    logger.info(f"수동 입금확인: order_id={order_id}")
    return jsonify({'ok': True})


@app.route('/api/orders/<int:order_id>/unconfirm', methods=['POST'])
def manual_unconfirm_order(order_id):
    """입금확인 취소 (대기로 되돌리기)"""
    conn = get_conn()
    conn.execute('''UPDATE orders SET status='pending', confirmed_at=NULL, pay_type=NULL
                   WHERE id=?''', (order_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/orders/by-buyer', methods=['DELETE'])
def delete_orders_by_buyer():
    """한 구매자의 (해당 세션) 주문 전체 삭제"""
    session_id = request.args.get('session_id')
    buyer = request.args.get('buyer_name')
    if not session_id or not buyer:
        return jsonify({'error': '필수 파라미터 없음'}), 400
    conn = get_conn()
    n = conn.execute("DELETE FROM orders WHERE session_id=? AND buyer_name=?", (session_id, buyer)).rowcount
    conn.commit(); conn.close()
    logger.info(f"🗑 주문 삭제: {buyer} (session {session_id}) {n}건")
    return jsonify({'ok': True, 'deleted': n})


@app.route('/api/purchase-history', methods=['GET'])
def purchase_history():
    """전체 구매 이력 (구매자/상품/단가/날짜). 닉네임·상품명 검색은 프론트에서 필터."""
    conn = get_conn()
    rows = conn.execute('''
        SELECT o.buyer_name, o.item, o.item_no, o.amount, o.status,
               ls.live_date, ls.filename
        FROM orders o LEFT JOIN live_sessions ls ON o.session_id = ls.id
        WHERE o.buyer_name NOT LIKE '%문자발송%'
          AND o.buyer_name NOT LIKE '%입금x%'
          AND o.buyer_name NOT LIKE '%입금 x%'
        ORDER BY o.buyer_name COLLATE NOCASE ASC, ls.live_date DESC, o.id ASC
    ''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/delivery/excel', methods=['GET'])
def download_delivery_excel():
    """입금확인 완료 엑셀 — 구매자별로 회원명단의 전화번호/주소 자동 매핑.
    구매자명을 키로 nick_mappings(realname)도 fallback 조회 → members.name 매칭.
    """
    session_id = request.args.get('session_id')
    conn = get_conn()

    # 구매자별 한 행 — 상품은 ',' 로 합치고, 금액은 SUM
    sql = ("SELECT o.buyer_name, GROUP_CONCAT(o.item, ', ') AS items, "
           "SUM(o.amount) AS total_amount, MIN(o.pay_type) AS pay_type, "
           "MAX(o.confirmed_at) AS confirmed_at, ls.live_date "
           "FROM orders o JOIN live_sessions ls ON o.session_id=ls.id "
           "WHERE o.status='confirmed'")
    params = []
    if session_id:
        sql += " AND o.session_id=?"
        params.append(session_id)
    sql += " GROUP BY o.session_id, o.buyer_name ORDER BY MAX(o.confirmed_at) DESC"

    items = conn.execute(sql, params).fetchall()

    # 회원명단 lookup: 이름 → row
    member_map = {}
    for m in conn.execute('SELECT * FROM members').fetchall():
        member_map[m['name']] = dict(m)
    # 닉네임 매핑(positive)으로 buyer_name → realname 변환
    nick_map = {}
    for n in conn.execute("SELECT nickname, realname FROM nick_mappings WHERE COALESCE(negative,0)=0").fetchall():
        if n['nickname'] and n['realname']:
            nick_map[n['nickname']] = n['realname']

    conn.close()

    def lookup_member(buyer):
        if not buyer: return None
        if buyer in member_map: return member_map[buyer]
        # 닉네임 매핑 따라가기
        real = nick_map.get(buyer)
        if real and real in member_map: return member_map[real]
        # 역방향 — buyer가 realname인 경우, 그 자체 이름으로 회원
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '입금확인 배송목록'

    headers = ['구매자명', '전화번호', '우편번호', '받는분주소(전체, 분할)', '배송메세지1',
               '상품', '금액', '결제방법', '입금확인일시', '라이브날짜']
    header_fill = PatternFill(fill_type='solid', fgColor='1F6B2E')
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)

    pay_labels = {'card': '카드결제', 'transfer': '계좌이체', 'manual': '수동확인', None: '', '': ''}
    matched = 0
    unmatched = 0
    for ri, row in enumerate(items, 2):
        buyer = row['buyer_name']
        m = lookup_member(buyer)
        if m: matched += 1
        else: unmatched += 1
        ws.cell(row=ri, column=1, value=buyer)
        ws.cell(row=ri, column=2, value=(m or {}).get('phone',''))
        ws.cell(row=ri, column=3, value=(m or {}).get('postal_code',''))
        ws.cell(row=ri, column=4, value=(m or {}).get('address',''))
        ws.cell(row=ri, column=5, value=(m or {}).get('message',''))
        ws.cell(row=ri, column=6, value=row['items'])
        ws.cell(row=ri, column=7, value=row['total_amount'])
        ws.cell(row=ri, column=8, value=pay_labels.get(row['pay_type'], row['pay_type'] or ''))
        ws.cell(row=ri, column=9, value=row['confirmed_at'])
        ws.cell(row=ri, column=10, value=row['live_date'])

    # 컬럼 폭 자동(대략)
    widths = [14, 15, 10, 50, 30, 60, 12, 10, 19, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w

    logger.info(f"📦 배송목록 다운로드: 총 {len(items)}명 (회원명단 매칭 {matched}, 미매칭 {unmatched})")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f'입금확인_배송목록_{now_kst().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════
#  회원명단 (이름 → 전화번호/주소 매핑)  ★ v17
# ══════════════════════════════════════════════════════════════════
@app.route('/api/members/list', methods=['GET'])
def members_list():
    conn = get_conn()
    rows = conn.execute('SELECT * FROM members ORDER BY name').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/members/upload', methods=['POST'])
def members_upload():
    """엑셀 업로드 — 헤더: 이름, 전화번호, 우편번호, 받는분주소(전체, 분할), 배송메세지1
    [v20] robust: .xls 지원, 다중 시트 자동 탐색, fuzzy 헤더 매칭."""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    fn = (file.filename or '').lower()
    # 모든 시트 후보 수집 (xlsx/xls 둘 다 지원)
    sheets_rows = []  # [(sheet_name, [rows])]
    try:
        if fn.endswith('.xls') and not fn.endswith('.xlsx'):
            import xlrd
            data = file.read()
            wb = xlrd.open_workbook(file_contents=data)
            for si in range(wb.nsheets):
                sh = wb.sheet_by_index(si)
                rr = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
                sheets_rows.append((sh.name, rr))
        else:
            wb = openpyxl.load_workbook(file, data_only=True)
            for sn in wb.sheetnames:
                rr = [list(row) for row in wb[sn].values]
                sheets_rows.append((sn, rr))
    except Exception as e:
        return jsonify({'error': f'엑셀 파싱 오류: {e}'}), 400

    # 시트별로 헤더 탐색
    def detect_header(rows):
        for i, row in enumerate(rows[:10]):
            if not row: continue
            cells = [str(c or '').strip() for c in row]
            nc = next((j for j, c in enumerate(cells) if '이름' in c or '성함' in c or '구매자' in c or '성명' in c or '수취인' in c or '수령인' in c), -1)
            pc = next((j for j, c in enumerate(cells) if '전화' in c or '핸드폰' in c or '폰번호' in c or 'phone' in c.lower() or '연락' in c), -1)
            oc = next((j for j, c in enumerate(cells) if '우편' in c or '우편번호' in c), -1)
            ac = next((j for j, c in enumerate(cells) if '주소' in c), -1)
            mc = next((j for j, c in enumerate(cells) if '메세지' in c or '메시지' in c or '메모' in c), -1)
            if nc >= 0 and (pc >= 0 or ac >= 0):
                return i, nc, pc, oc, ac, mc
        return None

    target = None
    for sn, rr in sheets_rows:
        det = detect_header(rr)
        if det:
            target = (sn, rr, det)
            break

    if not target:
        # 진단용: 어떤 시트에 어떤 헤더가 있었는지 알려주기
        diag = []
        for sn, rr in sheets_rows[:3]:
            head_preview = []
            for r in rr[:5]:
                cells = [str(c or '').strip() for c in (r or [])][:8]
                head_preview.append(' | '.join(cells))
            diag.append(f"[{sn}]: " + " / ".join(head_preview))
        logger.warning(f"회원명단 헤더 인식 실패 — 파일 구조: {' || '.join(diag)[:500]}")
        return jsonify({'error': '헤더(이름 + 전화번호/주소) 인식 실패. 첫 행에 "이름" 과 "전화번호" 또는 "주소" 키워드가 있어야 합니다.',
                        'sheets_found': [sn for sn, _ in sheets_rows]}), 400

    sn, rows, (header_idx, name_col, phone_col, post_col, addr_col, msg_col) = target
    logger.info(f"📇 회원명단 시트 '{sn}' 헤더={header_idx}행 name={name_col} phone={phone_col} post={post_col} addr={addr_col} msg={msg_col}")

    added = updated = skipped = 0
    conn = get_conn()
    try:
        for row in rows[header_idx + 1:]:
            if not row: continue
            row = list(row)
            name = str(row[name_col] or '').strip() if name_col < len(row) else ''
            if not name: skipped += 1; continue
            phone = str(row[phone_col] or '').strip() if phone_col >= 0 and phone_col < len(row) else ''
            post  = str(row[post_col]  or '').strip() if post_col  >= 0 and post_col  < len(row) else ''
            addr  = str(row[addr_col]  or '').strip() if addr_col  >= 0 and addr_col  < len(row) else ''
            msg   = str(row[msg_col]   or '').strip() if msg_col   >= 0 and msg_col   < len(row) else ''
            existing = conn.execute('SELECT id FROM members WHERE name=?', (name,)).fetchone()
            conn.execute(
                "INSERT OR REPLACE INTO members (name, phone, postal_code, address, message, updated_at) VALUES (?,?,?,?,?,?)",
                (name, phone, post, addr, msg, now_kst().isoformat())
            )
            if existing: updated += 1
            else: added += 1
        conn.commit()
    finally:
        conn.close()
    logger.info(f"📇 회원명단 import: 추가 {added}, 수정 {updated}, 스킵 {skipped}")
    return jsonify({'ok': True, 'added': added, 'updated': updated, 'skipped': skipped})


@app.route('/api/members/download', methods=['GET'])
def members_download():
    """회원명단 전체를 동일 양식으로 엑셀 다운로드"""
    conn = get_conn()
    rows = conn.execute('SELECT * FROM members ORDER BY name').fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '회원명단'
    headers = ['이름', '전화번호', '우편번호', '받는분주소(전체, 분할)', '배송메세지1']
    header_fill = PatternFill(fill_type='solid', fgColor='1F6B2E')
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r['name'])
        ws.cell(row=ri, column=2, value=r['phone'])
        ws.cell(row=ri, column=3, value=r['postal_code'])
        ws.cell(row=ri, column=4, value=r['address'])
        ws.cell(row=ri, column=5, value=r['message'])

    output = BytesIO(); wb.save(output); output.seek(0)
    fname = f'회원명단_{now_kst().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/members/<int:mid>', methods=['DELETE'])
def members_delete(mid):
    conn = get_conn()
    conn.execute('DELETE FROM members WHERE id=?', (mid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/members/<int:mid>', methods=['PUT', 'PATCH'])
def members_update(mid):
    """회원 정보 수정 (한 필드 또는 여러 필드 가능)"""
    data = request.get_json(silent=True) or {}
    fields = {}
    for k in ('name', 'phone', 'postal_code', 'address', 'message'):
        if k in data:
            fields[k] = str(data[k] or '').strip()
    if not fields:
        return jsonify({'error': '수정할 필드가 없습니다'}), 400
    conn = get_conn()
    try:
        if 'name' in fields:
            dup = conn.execute('SELECT id FROM members WHERE name=? AND id<>?', (fields['name'], mid)).fetchone()
            if dup:
                return jsonify({'error': f"이미 같은 이름의 회원이 있습니다: {fields['name']}"}), 400
        sets = ', '.join(f'{k}=?' for k in fields.keys()) + ', updated_at=?'
        params = list(fields.values()) + [now_kst().isoformat(), mid]
        conn.execute(f'UPDATE members SET {sets} WHERE id=?', params)
        conn.commit()
        row = conn.execute('SELECT * FROM members WHERE id=?', (mid,)).fetchone()
        return jsonify({'ok': True, 'member': dict(row) if row else None})
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════
#  총 판매품 리스트 (거래명세표 '판매품 리스트' 시트 백업)  ★ v17
# ══════════════════════════════════════════════════════════════════
@app.route('/api/products/list', methods=['GET'])
def products_list():
    session_id = request.args.get('session_id')
    conn = get_conn()
    sql = ("SELECT pc.*, ls.filename FROM product_catalog pc "
           "LEFT JOIN live_sessions ls ON pc.session_id=ls.id")
    if session_id:
        sql += " WHERE pc.session_id=? ORDER BY pc.id"
        rows = conn.execute(sql, (session_id,)).fetchall()
    else:
        sql += " ORDER BY pc.live_date DESC, pc.id"
        rows = conn.execute(sql).fetchall()
    conn.close()
    # 🌿 '하월시아 이름 + 판매가격'이 같은 중복 항목은 1건만 노출 (정렬 순서상 첫 건 유지)
    seen = set()
    deduped = []
    for r in rows:
        r = dict(r)
        key = ((r.get('item_name') or '').strip(), r.get('price'))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)
    return jsonify(deduped)


@app.route('/api/products/download', methods=['GET'])
def products_download():
    """판매품 카탈로그 전체(또는 한 세션) 엑셀 다운로드"""
    session_id = request.args.get('session_id')
    conn = get_conn()
    if session_id:
        rows = conn.execute('SELECT * FROM product_catalog WHERE session_id=? ORDER BY id', (session_id,)).fetchall()
    else:
        rows = conn.execute('SELECT * FROM product_catalog ORDER BY live_date DESC, id').fetchall()
    conn.close()

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '판매품 리스트'
    headers = ['라이브날짜', '번호', '하월시아 이름', '판매가격', '잔여갯수']
    header_fill = PatternFill(fill_type='solid', fgColor='1F6B2E')
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r['live_date'])
        ws.cell(row=ri, column=2, value=r['item_no'])
        ws.cell(row=ri, column=3, value=r['item_name'])
        ws.cell(row=ri, column=4, value=r['price'])
        ws.cell(row=ri, column=5, value=r['remaining'])

    output = BytesIO(); wb.save(output); output.seek(0)
    fname = f'판매품리스트_{now_kst().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/products/<int:pid>', methods=['PUT', 'PATCH'])
def products_update(pid):
    """판매품 한 건 수정 (이름/가격/잔여/번호)"""
    data = request.get_json(silent=True) or {}
    fields = {}
    if 'item_name' in data: fields['item_name'] = str(data['item_name'] or '').strip()
    if 'price' in data:
        try:
            fields['price'] = int(data['price']) if data['price'] not in (None, '') else None
        except Exception:
            return jsonify({'error': '가격이 숫자가 아닙니다'}), 400
    if 'remaining' in data: fields['remaining'] = str(data['remaining'] or '').strip()
    if 'item_no' in data: fields['item_no'] = str(data['item_no'] or '').strip()
    if not fields:
        return jsonify({'error': '수정할 필드가 없습니다'}), 400
    conn = get_conn()
    try:
        sets = ', '.join(f'{k}=?' for k in fields.keys())
        params = list(fields.values()) + [pid]
        conn.execute(f'UPDATE product_catalog SET {sets} WHERE id=?', params)
        conn.commit()
        row = conn.execute('SELECT * FROM product_catalog WHERE id=?', (pid,)).fetchone()
        return jsonify({'ok': True, 'product': dict(row) if row else None})
    finally:
        conn.close()


@app.route('/api/products/cleanup', methods=['POST'])
def products_cleanup():
    """50만원 이하(또는 가격없음) 판매품 삭제 + 같은 이름 중복 제거(이름당 최신 1건만 유지)."""
    conn = get_conn()
    try:
        removed_low = conn.execute(
            "DELETE FROM product_catalog WHERE price IS NULL OR price <= ?",
            (CATALOG_MIN_PRICE,)
        ).rowcount
        removed_dup = conn.execute(
            "DELETE FROM product_catalog WHERE id NOT IN "
            "(SELECT MAX(id) FROM product_catalog GROUP BY item_name)"
        ).rowcount
        conn.commit()
        remaining = conn.execute("SELECT COUNT(*) AS c FROM product_catalog").fetchone()['c']
        logger.info(f"🧹 판매품 정리: 저가 {removed_low} + 중복 {removed_dup} 삭제, 남은 {remaining}")
        return jsonify({'ok': True, 'removed_low': removed_low, 'removed_dup': removed_dup, 'remaining': remaining})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/products/<int:pid>', methods=['DELETE'])
def products_delete(pid):
    conn = get_conn()
    conn.execute('DELETE FROM product_catalog WHERE id=?', (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/products/upload', methods=['POST'])
def products_upload():
    """판매품 리스트 수동 업로드 (백업/복원).
       헤더: [라이브날짜], 번호, 하월시아 이름, 판매가격, 잔여갯수
       기존 백업본(session_id=0/NULL)은 모두 삭제하고 새로 채움."""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = None
        for sn in wb.sheetnames:
            if '판매품' in sn or '재고' in sn:
                ws = wb[sn]; break
        if ws is None:
            ws = wb.active
        rows_raw = list(ws.values)
    except Exception as e:
        return jsonify({'error': f'엑셀 파싱 오류: {e}'}), 400

    # 헤더 찾기
    header_idx = no_col = name_col = price_col = rem_col = date_col = -1
    for i, row in enumerate(rows_raw[:8]):
        if not row: continue
        cells = [str(c or '').strip() for c in row]
        ni = next((j for j, c in enumerate(cells) if c == '번호' or '연번' in c or c.lower() == 'no'), -1)
        nn = next((j for j, c in enumerate(cells) if '하월시아' in c or '품명' in c or '상품' in c or c == '이름'), -1)
        pp = next((j for j, c in enumerate(cells) if '가격' in c or '금액' in c), -1)
        rr = next((j for j, c in enumerate(cells) if '잔여' in c or '재고' in c or '수량' in c), -1)
        dc = next((j for j, c in enumerate(cells) if '라이브' in c or '날짜' in c), -1)
        if ni >= 0 and nn >= 0 and pp >= 0:
            header_idx = i; no_col=ni; name_col=nn; price_col=pp; rem_col=rr; date_col=dc
            break
    if header_idx < 0:
        return jsonify({'error': '헤더(번호/이름/가격) 인식 실패'}), 400

    added = 0
    conn = get_conn()
    try:
        conn.execute("DELETE FROM product_catalog WHERE session_id IS NULL OR session_id=0")
        seen = set(r[0] for r in conn.execute("SELECT DISTINCT item_name FROM product_catalog").fetchall())
        for row in rows_raw[header_idx+1:]:
            if not row: continue
            row = list(row)
            no_v = row[no_col] if no_col < len(row) else None
            nm_v = row[name_col] if name_col < len(row) else None
            pr_v = row[price_col] if price_col < len(row) else None
            rm_v = row[rem_col] if rem_col >= 0 and rem_col < len(row) else None
            dt_v = row[date_col] if date_col >= 0 and date_col < len(row) else None
            if (no_v is None or str(no_v).strip() == '') and (not nm_v or not str(nm_v).strip()):
                continue
            if isinstance(no_v, (int, float)) and float(no_v).is_integer():
                item_no = str(int(no_v))
            else:
                item_no = str(no_v).strip() if no_v is not None else ''
            try:
                raw = str(pr_v or '').replace(',', '').replace('원', '').strip()
                price = int(float(raw)) if raw else None
            except Exception:
                price = None
            item_name = str(nm_v or '').strip()
            remaining = str(rm_v).strip() if rm_v is not None and str(rm_v).strip() else ''
            live_date = str(dt_v).strip() if dt_v else ''
            if not _catalog_keep(price):   # 50만원 이하/가격없음 제외
                continue
            if item_name in seen:          # 같은 이름 중복 제외
                continue
            seen.add(item_name)
            conn.execute(
                "INSERT INTO product_catalog (session_id, live_date, item_no, item_name, price, remaining) VALUES (?,?,?,?,?,?)",
                (0, live_date, item_no, item_name, price, remaining)
            )
            added += 1
        conn.commit()
    finally:
        conn.close()
    logger.info(f"📚 판매품 리스트 import: {added}건")
    return jsonify({'ok': True, 'added': added})



# ══════════════════════════════════════════════════════════════════
#  위탁판매 정산 (★ v24)
#  거래명세표 item에 "-XXX위탁" suffix가 있는 confirmed 주문만 대상.
#  - XXX 부분을 위탁자 ID로 추출
#  - 타업체 계수(0.9 or 1.0) × 라방계수(0.8) → 라방플랫폼수수료 제외 가격
#  - × 0.8 → 송금 정산 시 가격
# ══════════════════════════════════════════════════════════════════
import re as _re_consign

def _is_consign(item_str):
    """상품명에 '위탁'이라는 단어가 들어가면 위탁품으로 인식.
    예) '옵투사 실생(이계)-SSAC위탁', '페일피스 -IB위탁품', '단정(오리지널)-IB위탁품' 모두 True.
    이름은 가공하지 않고 거래명세표에 올린 원본 그대로 사용한다."""
    if not item_str:
        return False
    return '위탁' in str(item_str)


def _consign_name(item_str):
    """정산표 '이름' 칸에 들어갈 값 = 거래명세표 원본 그대로(앞뒤 공백만 정리)."""
    return str(item_str or '').strip()


# 위탁자 코드 추출 — 상품명에서 '위탁' 바로 앞 코드(SSAC/IB/SW 등).
#   예) '옵투사 실생(이계)-SSAC위탁' → SSAC
#       '페일피스 -IB위탁품'        → IB
#       '단정(오리지널)-IB위탁품'    → IB
_CONSIGNOR_RE = _re_consign.compile(r'([0-9A-Za-z가-힣]+)\s*위탁')

def _consignor_of(item_str):
    """상품명에서 위탁자 코드를 추출. '위탁'은 있으나 코드 패턴이 없으면 '기타',
    아예 위탁품이 아니면 None."""
    if not item_str:
        return None
    s = str(item_str)
    if '위탁' not in s:
        return None
    matches = _CONSIGNOR_RE.findall(s)
    if matches:
        return matches[-1].strip()   # 마지막 '위탁' 바로 앞 코드 사용
    return '기타'


@app.route('/api/consignment/list', methods=['GET'])
def consignment_list():
    """위탁자(코드)별 위탁판매 정산 데이터.
    · 상품명에 '위탁'이 들어간 모든 주문(상태 무관: 대기 포함)을 집계.
    · 상품명 끝의 위탁자 코드(-SSAC위탁 → SSAC, -IB위탁품 → IB)별로 묶어 코드마다 카드 1개.
    session_id 옵션. 각 row: order_id, item_name(원본 그대로), price, qty,
       is_3rd_party, sale_total, after_platform, after_remit
    """
    session_id = request.args.get('session_id')
    conn = get_conn()
    sql = ("SELECT o.id, o.buyer_name, o.item, o.amount, o.session_id, o.is_3rd_party, "
           "       ls.live_date, ls.filename "
           "FROM orders o LEFT JOIN live_sessions ls ON o.session_id=ls.id "
           "WHERE 1=1")
    params = []
    if session_id:
        sql += " AND o.session_id=?"
        params.append(session_id)
    sql += " ORDER BY o.id ASC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    grouped = {}
    seq = {}
    for r in rows:
        cons = _consignor_of(r['item'])
        if not cons:
            continue
        name = _consign_name(r['item'])
        price = int(r['amount'] or 0)
        qty = 1
        sale_total = price * qty
        is_3p = 1 if (r['is_3rd_party'] or 0) else 0
        third_factor = 0.9 if is_3p else 1.0
        live_factor = 0.8
        after_platform = round(sale_total * third_factor * live_factor)
        after_remit = round(after_platform * 0.8)

        if cons not in grouped:
            grouped[cons] = {
                'consignor': cons,
                'live_date': r['live_date'],
                'filename': r['filename'],
                'rows': [],
                'sum_sale': 0, 'sum_after_platform': 0, 'sum_after_remit': 0,
            }
            seq[cons] = 0
        seq[cons] += 1
        grouped[cons]['rows'].append({
            'no': seq[cons],
            'order_id': r['id'],
            'item_name': name,
            'item_full': r['item'],
            'buyer_name': (r['buyer_name'] or '').strip(),
            'price': price,
            'qty': qty,
            'sale_total': sale_total,
            'is_3rd_party': is_3p,
            'third_factor': third_factor,
            'live_factor': live_factor,
            'after_platform': after_platform,
            'after_remit': after_remit,
            'live_date': r['live_date'],
        })
        grouped[cons]['sum_sale'] += sale_total
        grouped[cons]['sum_after_platform'] += after_platform
        grouped[cons]['sum_after_remit'] += after_remit

    # 정렬: 위탁자 코드 가나다/영문순
    return jsonify(sorted(grouped.values(), key=lambda x: x['consignor']))


@app.route('/api/orders/<int:order_id>/3rd-party', methods=['PUT'])
def set_3rd_party(order_id):
    """타업체 제품 유무 토글"""
    data = request.get_json(silent=True) or {}
    flag = 1 if data.get('is_3rd_party') else 0
    conn = get_conn()
    try:
        conn.execute("UPDATE orders SET is_3rd_party=? WHERE id=?", (flag, order_id))
        conn.commit()
        return jsonify({'ok': True, 'is_3rd_party': flag})
    finally:
        conn.close()


@app.route('/api/consignment/excel', methods=['GET'])
def consignment_excel():
    """위탁자(코드) 한 명의 위탁 정산표를 엑셀로 다운로드.
    · 상품명에 '위탁'이 들어간 모든 주문(상태 무관)을 집계.
    · 상품명 끝의 위탁자 코드(-SSAC위탁 → SSAC)별로 묶음.
    · 제목 = '{정산날짜 M월 D일} {업로드 파일명(확장자 제외)} / 유튜브'
    · 정산날짜 = 해당 세션 날짜, 이름 = 거래명세표 원본 그대로.
    · 모든 폰트 맑은 고딕, 기존 대비 1pt 상향.
    layout: 정산날짜 | 번호 | 이름 | 가격 | 수량 | 판매가 합 | 타업체(10%) | 라방판매(20%)
            | 라방플랫폼수수료 제외 가격 | 송금 정산 시 가격 (20%제)
    """
    # consignor(신규) 우선, 과거 호환을 위해 buyer 파라미터도 허용
    consignor = (request.args.get('consignor') or request.args.get('buyer') or '').strip()
    session_id = request.args.get('session_id')
    if not consignor:
        return jsonify({'error': 'consignor 필요'}), 400

    conn = get_conn()
    sql = ("SELECT o.id, o.buyer_name, o.item, o.amount, o.is_3rd_party, "
           "       ls.live_date, ls.filename "
           "FROM orders o LEFT JOIN live_sessions ls ON o.session_id=ls.id "
           "WHERE 1=1")
    params = []
    if session_id:
        sql += " AND o.session_id=?"
        params.append(session_id)
    sql += " ORDER BY o.id ASC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    items = []        # (order_id, name, price, live_date, is_3rd_party)
    src_filename = '' # 제목에 쓸 업로드 파일명
    title_date = ''   # 제목/정산날짜용 세션 날짜
    for r in rows:
        if _consignor_of(r['item']) != consignor:
            continue
        if not src_filename:
            src_filename = r['filename'] or ''
        if not title_date:
            title_date = r['live_date'] or ''
        items.append((r['id'], _consign_name(r['item']), int(r['amount'] or 0),
                      r['live_date'], 1 if (r['is_3rd_party'] or 0) else 0))

    FONT = '맑은 고딕'
    # 폰트 크기: 기존 대비 1pt 상향 (제목 13→14, 그 외 11→12)
    SZ_TITLE, SZ_BODY = 14, 12

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (consignor[:25] + ' 정산')

    # ── 제목: '{M월 D일} {파일명(확장자 제외)} / 유튜브' ──
    fname_noext = os.path.splitext(src_filename)[0] if src_filename else ''
    try:
        _d = datetime.strptime(title_date, '%Y-%m-%d')
        date_kr = f"{_d.month}월 {_d.day}일"
    except Exception:
        date_kr = title_date or ''
    title_text = " ".join(p for p in [date_kr, fname_noext] if p).strip()
    if title_text:
        title_text += " / 유튜브"
    else:
        title_text = "유튜브"
    tcell = ws.cell(row=1, column=1, value=title_text)
    tcell.font = Font(name=FONT, bold=True, size=SZ_TITLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    # ── 헤더 ──
    headers = ['정산\n날짜', '번호', '이름', '가격', '수량', '판매가 합',
               '타업체\n제품 유무(10%)', '라방판매(20%)',
               '라방플랫폼수수료 제외 가격\n(적립금 100% 사용가능)',
               '송금 정산 시\n가격 (20% 제)']
    header_fill   = PatternFill(fill_type='solid', fgColor='F2F2F2')
    white_fill    = PatternFill(fill_type='solid', fgColor='FFFFFF')
    orange_fill   = PatternFill(fill_type='solid', fgColor='FFA500')
    green_fill    = PatternFill(fill_type='solid', fgColor='9BE6A1')
    yellow_fill   = PatternFill(fill_type='solid', fgColor='FFFF66')

    from openpyxl.styles import Alignment, Border, Side
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name=FONT, bold=True, size=SZ_BODY)
        cell.alignment = center
        cell.border = border
        if ci == 7:    cell.fill = orange_fill          # 타업체 제품 유무
        elif ci in (9, 10): cell.fill = green_fill      # 라방플랫폼/송금
        else:          cell.fill = header_fill

    # ── 데이터 ──
    sum_after_platform = 0
    sum_after_remit = 0
    for i, (oid, name, price, dt, is_3p) in enumerate(items, 1):
        try:
            d = datetime.strptime(dt or '', '%Y-%m-%d')
            dkr = f"{d.month:02d}월 {d.day:02d}일"
        except Exception:
            dkr = dt or ''
        third = 0.9 if is_3p else 1.0
        sale = price * 1
        after_p = round(sale * third * 0.8)
        after_r = round(after_p * 0.8)
        sum_after_platform += after_p
        sum_after_remit += after_r
        row = i + 2
        cells = [
            (1, dkr,    white_fill),
            (2, i,      white_fill),
            (3, name,   white_fill),
            (4, price,  white_fill),
            (5, 1,      white_fill),
            (6, sale,   white_fill),
            (7, third,  orange_fill),
            (8, 0.8,    white_fill),
            (9, after_p, green_fill),
            (10, after_r, green_fill),
        ]
        for ci, val, fill in cells:
            c = ws.cell(row=row, column=ci, value=val)
            if ci == 3:
                c.alignment = left
            elif ci in (4, 5, 6, 9, 10):
                c.alignment = right
            else:
                c.alignment = center
            c.border = border
            c.fill = fill
            c.font = Font(name=FONT, size=SZ_BODY)
            if isinstance(val, int) and ci in (4, 6, 9, 10):
                c.number_format = '#,##0'

    # ── 합계 행 (마지막) ──
    bottom_row = len(items) + 3
    for ci, val in ((9, sum_after_platform), (10, sum_after_remit)):
        c = ws.cell(row=bottom_row, column=ci, value=val)
        c.fill = yellow_fill
        c.number_format = '#,##0'
        c.font = Font(name=FONT, bold=True, size=SZ_BODY)
        c.alignment = right
        c.border = border

    # ── 컬럼폭 ──
    from openpyxl.utils import get_column_letter
    widths = [10, 8, 40, 12, 8, 12, 12, 11, 20, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 40

    output = BytesIO()
    wb.save(output); output.seek(0)
    fname = f'위탁정산_{consignor}_{now_kst().strftime("%Y%m%d")}.xlsx'
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════
#  닉네임 매핑
# ══════════════════════════════════════════════════════════════════
@app.route('/api/mappings', methods=['GET'])
def get_mappings():
    conn = get_conn()
    rows = conn.execute('SELECT * FROM nick_mappings ORDER BY nickname').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/mappings', methods=['POST'])
def add_mapping():
    data = request.get_json(force=True) or {}
    conn = get_conn()
    conn.execute('INSERT OR REPLACE INTO nick_mappings (nickname, realname) VALUES (?, ?)',
                 (data['nickname'], data['realname']))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/mappings/<nickname>', methods=['DELETE'])
def delete_mapping(nickname):
    conn = get_conn()
    conn.execute('DELETE FROM nick_mappings WHERE nickname=?', (nickname,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════
#  거래명세표 발송 진행상태 (식물보관 / 배송완료)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/buyer/status', methods=['POST'])
def set_buyer_status():
    """구매자별 발송 상태 변경. body: {session_id, buyer_name, status}
       status: 'stored' (식물보관) | 'shipped' (배송완료) | 'reset' (되돌리기/null)"""
    data = request.get_json(silent=True) or {}
    sid = data.get('session_id')
    buyer = (data.get('buyer_name') or '').strip()
    status = (data.get('status') or '').strip()
    if not sid or not buyer:
        return jsonify({'error': 'session_id, buyer_name 필요'}), 400
    if status not in ('stored', 'shipped', 'reset'):
        return jsonify({'error': "status는 'stored', 'shipped', 'reset' 중 하나"}), 400
    conn = get_conn()
    try:
        if status == 'reset':
            conn.execute("DELETE FROM buyer_status WHERE session_id=? AND buyer_name=?", (sid, buyer))
        else:
            conn.execute(
                "INSERT OR REPLACE INTO buyer_status (session_id, buyer_name, status, updated_at) VALUES (?,?,?,?)",
                (sid, buyer, status, now_kst().isoformat())
            )
        conn.commit()
        logger.info(f"  📦 구매자 상태 변경: session={sid} '{buyer}' → {status}")
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/buyer/status/export', methods=['GET'])
def export_buyer_status():
    """식물 보관 + 배송완료 상태 전체를 CSV로 다운로드 (서버 초기화 대비)"""
    import csv
    from io import StringIO
    from flask import Response
    conn = get_conn()
    rows = conn.execute(
        "SELECT bs.session_id, bs.buyer_name, bs.status, bs.updated_at, ls.live_date, ls.filename "
        "FROM buyer_status bs LEFT JOIN live_sessions ls ON ls.id = bs.session_id "
        "ORDER BY bs.session_id, bs.buyer_name"
    ).fetchall()
    conn.close()
    buf = StringIO()
    buf.write('\ufeff')  # UTF-8 BOM
    w = csv.writer(buf)
    w.writerow(['session_id', 'buyer_name', 'status', 'updated_at', 'live_date', 'filename'])
    for r in rows:
        w.writerow([r['session_id'], r['buyer_name'], r['status'], r['updated_at'], r['live_date'] or '', r['filename'] or ''])
    fname = f'buyer_status_{now_kst().strftime("%Y%m%d_%H%M")}.csv'
    return Response(buf.getvalue(),
                    mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/buyer/status/import', methods=['POST'])
def import_buyer_status():
    """식물 보관 CSV 일괄 복원 (UPSERT)"""
    import csv
    from io import StringIO
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    try:
        raw = file.read()
        text = None
        for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
            try: text = raw.decode(enc); break
            except UnicodeDecodeError: continue
        if text is None:
            return jsonify({'error': '파일 인코딩 해석 실패'}), 400
    except Exception as e:
        return jsonify({'error': f'파일 읽기 오류: {e}'}), 400

    added = skipped = 0
    conn = get_conn()
    try:
        reader = csv.DictReader(StringIO(text))
        fns = [(fn or '').strip().lower() for fn in (reader.fieldnames or [])]
        if 'session_id' not in fns or 'buyer_name' not in fns or 'status' not in fns:
            return jsonify({'error': "CSV에 'session_id,buyer_name,status' 헤더 필요"}), 400
        for row in csv.DictReader(StringIO(text)):
            sid = (row.get('session_id') or '').strip()
            buyer = (row.get('buyer_name') or '').strip()
            status = (row.get('status') or '').strip()
            updated = (row.get('updated_at') or now_kst().isoformat()).strip()
            if not sid or not buyer or status not in ('stored', 'shipped'):
                skipped += 1; continue
            try:
                conn.execute(
                    "INSERT OR REPLACE INTO buyer_status (session_id, buyer_name, status, updated_at) VALUES (?,?,?,?)",
                    (int(sid), buyer, status, updated)
                )
                added += 1
            except Exception as e:
                skipped += 1
                logger.warning(f"buyer_status import 스킵: {e}")
        conn.commit()
    finally:
        conn.close()
    logger.info(f"📤 buyer_status import: 추가/수정 {added}, 스킵 {skipped}")
    return jsonify({'ok': True, 'added': added, 'skipped': skipped})


@app.route('/api/buyer/status', methods=['GET'])
def get_buyer_status():
    """구매자 상태 조회.
       session_id 있으면 {buyer_name: status} 매핑 반환
       session_id 없으면 [{session_id, buyer_name, status}, ...] 리스트 반환 (v23)"""
    sid = request.args.get('session_id')
    conn = get_conn()
    if sid:
        rows = conn.execute(
            "SELECT buyer_name, status FROM buyer_status WHERE session_id=?",
            (sid,)
        ).fetchall()
        conn.close()
        return jsonify({r['buyer_name']: r['status'] for r in rows})
    else:
        rows = conn.execute("SELECT session_id, buyer_name, status FROM buyer_status").fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════
#  수동 확인 실행 버튼
# ══════════════════════════════════════════════════════════════════
@app.route('/api/check/run', methods=['POST'])
def manual_check():
    run_auto_check()
    return jsonify({'ok': True, 'message': '입금확인 완료'})


# ══════════════════════════════════════════════════════════════════
#  7일 자동확인 진행현황 조회  ⭐ (이전 버전에서 데코레이터 누락이 있었음)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/check/logs', methods=['GET'])
def get_check_logs():
    """세션별 7일 자동확인 실행 로그"""
    session_id = request.args.get('session_id')
    conn = get_conn()

    # 세션 정보 조회
    if session_id:
        session = conn.execute('SELECT * FROM live_sessions WHERE id=?', (session_id,)).fetchone()
    else:
        session = conn.execute('SELECT * FROM live_sessions ORDER BY created_at DESC LIMIT 1').fetchone()

    if not session:
        conn.close()
        return jsonify({'session': None, 'days': []})

    session = dict(session)

    # 7일 날짜 생성 — 라방날짜 +1일부터 시작 (check_start 미설정/구버전 세션도 안전)
    from datetime import date
    try:
        start = datetime.strptime(session['live_date'], '%Y-%m-%d').date() + timedelta(days=1)
    except Exception:
        start = datetime.strptime(session['check_start'], '%Y-%m-%d').date()
    days = []
    for i in range(7):
        d = start + timedelta(days=i)
        log = conn.execute(
            'SELECT * FROM check_logs WHERE session_id=? AND check_date=? ORDER BY id DESC LIMIT 1',
            (session['id'], str(d))
        ).fetchone()
        today = now_kst().date()
        if d < today:
            day_status = 'done' if log else 'missed'
        elif d == today:
            day_status = 'today'
        else:
            day_status = 'future'

        days.append({
            'date': str(d),
            'day_status': day_status,
            'log': dict(log) if log else None
        })

    conn.close()
    return jsonify({'session': session, 'days': days})


# ══════════════════════════════════════════════════════════════════
#  SMS 수신 목록 (모니터링용)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/sms/recent', methods=['GET'])
def recent_sms():
    conn = get_conn()
    rows = conn.execute('''SELECT * FROM sms_payments
                           ORDER BY received_at DESC LIMIT 30''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════════
#  상태 요약
# ══════════════════════════════════════════════════════════════════
@app.route('/api/myip', methods=['GET'])
def get_my_ip():
    """이 PC의 로컬 IP 자동 감지"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
    except Exception:
        ip = "127.0.0.1"
    return jsonify({'ip': ip, 'sms_url': f'http://{ip}:5000/sms'})


@app.route('/api/status', methods=['GET'])
def get_status():
    conn = get_conn()
    total     = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    confirmed = conn.execute("SELECT COUNT(*) FROM orders WHERE status='confirmed'").fetchone()[0]
    pending   = conn.execute("SELECT COUNT(*) FROM orders WHERE status='pending'").fetchone()[0]
    sms_today = conn.execute("SELECT COUNT(*) FROM sms_payments WHERE DATE(received_at)=DATE('now')").fetchone()[0]
    candidates_open = conn.execute("SELECT COUNT(*) FROM match_candidates WHERE status='open'").fetchone()[0]
    conn.close()
    return jsonify({'total': total, 'confirmed': confirmed, 'pending': pending,
                    'sms_today': sms_today, 'candidates_open': candidates_open})


# ══════════════════════════════════════════════════════════════════
#  의심후보 조회 / 승인 / 거절
# ══════════════════════════════════════════════════════════════════
@app.route('/api/candidates', methods=['GET'])
def list_candidates():
    session_id = request.args.get('session_id')
    status     = request.args.get('status', 'open')
    conn = get_conn()
    sql = """SELECT mc.*, ls.live_date, ls.filename
             FROM match_candidates mc
             LEFT JOIN live_sessions ls ON mc.session_id=ls.id
             WHERE mc.status=?"""
    params = [status]
    if session_id:
        sql += ' AND mc.session_id=?'
        params.append(session_id)
    sql += """ ORDER BY CASE mc.confidence WHEN 'high' THEN 0 WHEN 'medium' THEN 1
                          WHEN 'low' THEN 2 ELSE 3 END, mc.created_at DESC"""
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/candidates/<int:cand_id>/approve', methods=['POST'])
def approve_candidate(cand_id):
    """승인 — 같은 buyer_name + session의 모든 pending 행을 일괄 confirmed."""
    data = request.get_json(silent=True) or {}
    override_buyer = (data.get('buyer_name') or '').strip()
    conn = get_conn()
    try:
        cand = conn.execute('SELECT * FROM match_candidates WHERE id=?', (cand_id,)).fetchone()
        if not cand: return jsonify({'error': '후보 없음'}), 404
        cand = dict(cand)
        target_buyer = override_buyer or cand.get('candidate_buyer_name')
        target_session = cand.get('session_id')
        if not target_buyer:
            return jsonify({'error': '구매자 이름 필요'}), 400
        pending = conn.execute(
            "SELECT id, amount FROM orders WHERE session_id=? AND buyer_name=? AND status='pending'",
            (target_session, target_buyer)).fetchall()
        if not pending:
            # 이 buyer가 다른 경로(이전 후보 승인, 자동매칭 등)로 이미 confirmed 됐을 가능성 체크
            any_orders = conn.execute(
                "SELECT id, status FROM orders WHERE session_id=? AND buyer_name=?",
                (target_session, target_buyer)).fetchall()
            if any_orders:
                # 이미 입금확인된 구매자 — 후보만 닫고 매핑은 추가
                now = now_kst().isoformat()
                conn.execute("UPDATE match_candidates SET status='approved', decided_at=?, candidate_buyer_name=? WHERE id=?",
                             (now, target_buyer, cand_id))
                paid_name = cand.get('paid_name')
                if paid_name and target_buyer and paid_name != target_buyer:
                    try:
                        conn.execute(
                            "INSERT OR REPLACE INTO nick_mappings (nickname, realname, negative) VALUES (?, ?, 0)",
                            (paid_name, target_buyer))
                        logger.info(f"  🔗 매핑 자동 추가 (이미 확인): '{paid_name}' ↔ '{target_buyer}'")
                    except Exception as e:
                        logger.warning(f"매핑 자동추가 실패: {e}")
                if cand['source'] == 'sms' and cand.get('source_ref'):
                    try: conn.execute('UPDATE sms_payments SET matched=1 WHERE id=?', (int(cand['source_ref']),))
                    except: pass
                conn.commit()
                logger.info(f"  ℹ️ 이미 확인된 구매자: '{target_buyer}' — 후보 닫고 매핑만 추가")
                return jsonify({'ok': True, 'buyer_name': target_buyer, 'rows_confirmed': 0, 'total_amount': 0,
                                'message': f"'{target_buyer}' 는 이미 입금확인됨 — 후보 닫고 매핑 추가"})
            return jsonify({'error': f"세션{target_session} '{target_buyer}' 주문이 없음 (이미 삭제됐을 수 있음)"}), 404
        pay_type = 'card' if cand['source'] == 'imweb' else 'transfer'
        now = now_kst().isoformat()
        for row in pending:
            conn.execute("UPDATE orders SET status='confirmed', confirmed_at=?, pay_type=?, bank_date=? WHERE id=?",
                         (now, pay_type, cand.get('paid_at'), row['id']))
        conn.execute("UPDATE match_candidates SET status='approved', decided_at=?, candidate_buyer_name=? WHERE id=?",
                     (now, target_buyer, cand_id))
        if cand['source'] == 'sms' and cand.get('source_ref'):
            try: conn.execute('UPDATE sms_payments SET matched=1 WHERE id=?', (int(cand['source_ref']),))
            except: pass
        # ⭐ 닉네임 매핑 자동 추가 (동일인) — 받은 이름과 구매자명이 다를 때만 의미있음
        paid_name = cand.get('paid_name')
        if paid_name and target_buyer and paid_name != target_buyer:
            try:
                conn.execute(
                    "INSERT OR REPLACE INTO nick_mappings (nickname, realname, negative) VALUES (?, ?, 0)",
                    (paid_name, target_buyer)
                )
                logger.info(f"  🔗 매핑 자동 추가 (동일인): '{paid_name}' ↔ '{target_buyer}'")
            except Exception as e:
                logger.warning(f"매핑 자동추가 실패: {e}")
        conn.commit()
        total = sum(r['amount'] for r in pending)
        logger.info(f"  👍 후보 승인: cand={cand_id} → '{target_buyer}' ({len(pending)}건 합{total:,}원)")
        return jsonify({'ok': True, 'buyer_name': target_buyer, 'rows_confirmed': len(pending), 'total_amount': total})
    finally:
        conn.close()


@app.route('/api/candidates/<int:cand_id>/delete', methods=['POST'])
def delete_candidate(cand_id):
    """삭제 — 후보를 그냥 제거. 닉네임 매핑 동일인/불일치 어느쪽도 추가하지 않음."""
    conn = get_conn()
    try:
        cand = conn.execute('SELECT * FROM match_candidates WHERE id=?', (cand_id,)).fetchone()
        if not cand:
            return jsonify({'error': '후보 없음'}), 404
        conn.execute("DELETE FROM match_candidates WHERE id=?", (cand_id,))
        conn.commit()
        logger.info(f"  🗑 후보 삭제: cand={cand_id} (매핑 변동 없음)")
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/candidates/<int:cand_id>/reject', methods=['POST'])
def reject_candidate(cand_id):
    """거절 — 후보 닫기 + 닉네임 매핑에 '동일인 아님(negative=1)' 자동 추가"""
    conn = get_conn()
    try:
        cand = conn.execute('SELECT * FROM match_candidates WHERE id=?', (cand_id,)).fetchone()
        if not cand:
            return jsonify({'error': '후보 없음'}), 404
        cand = dict(cand)
        conn.execute("UPDATE match_candidates SET status='rejected', decided_at=? WHERE id=?",
                     (now_kst().isoformat(), cand_id))
        # ⭐ 닉네임 매핑 자동 추가 (동일인 아님) — 같은 pair 다시 후보로 만들지 않게
        paid_name  = cand.get('paid_name')
        buyer_name = cand.get('candidate_buyer_name')
        if paid_name and buyer_name and paid_name != buyer_name:
            try:
                conn.execute(
                    "INSERT OR REPLACE INTO nick_mappings (nickname, realname, negative) VALUES (?, ?, 1)",
                    (paid_name, buyer_name)
                )
                logger.info(f"  🚫 매핑 자동 추가 (동일인 아님): '{paid_name}' ↔ '{buyer_name}'")
            except Exception as e:
                logger.warning(f"매핑 자동추가 실패: {e}")
        conn.commit()
        return jsonify({'ok': True})
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════
#  ★ SMS 진단 엔드포인트 (Automate 디버깅)
# ══════════════════════════════════════════════════════════════════
@app.route('/api/sms/echo', methods=['GET', 'POST'])
def sms_echo():
    info = {
        'ok': True, 'received_at': now_kst().isoformat(),
        'method': request.method, 'path': request.path,
        'remote_addr': request.headers.get('X-Forwarded-For', request.remote_addr),
        'content_type': request.content_type,
        'content_length': request.content_length,
        'headers': {k: v for k, v in request.headers.items()},
        'args': request.args.to_dict(flat=False),
        'form': request.form.to_dict(flat=False),
        'json': request.get_json(silent=True),
        'body_preview': (request.get_data(as_text=True) or '')[:500],
    }
    logger.info(f"🩺 /api/sms/echo {request.method}")
    return jsonify(info)


@app.route('/api/sms/simulate', methods=['GET', 'POST'])
def sms_simulate():
    body = (request.args.get('body') or request.form.get('body')
            or (request.get_json(silent=True) or {}).get('body') or '').strip()
    if not body:
        return jsonify({'ok': False, 'error': 'body 필요'}), 400
    parsed = parse_sms(body)
    recv_iso = now_kst().isoformat()
    sms_id = None
    try:
        conn = get_conn()
        cur = conn.execute("INSERT INTO sms_payments (sender, body, parsed_name, parsed_amount, received_at) VALUES (?,?,?,?,?)",
                            ('SIMULATE', body,
                             parsed['name']   if parsed else None,
                             parsed['amount'] if parsed else None,
                             recv_iso))
        sms_id = cur.lastrowid
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"simulate 저장: {e}")
    if parsed:
        match_sms_to_order(parsed, recv_iso, sms_id=sms_id)
    return jsonify({'ok': True, 'parsed': parsed, 'sms_id': sms_id})



# ══════════════════════════════════════════════════════════════════
#  은행 입금내역 엑셀 업로드 (수동 일괄 매칭)
# ══════════════════════════════════════════════════════════════════
def parse_bank_excel(file_storage):
    """KB/농협 등 은행 입금내역 엑셀(.xls/.xlsx)에서 (이름, 입금액, 일시) 추출"""
    fn = (file_storage.filename or '').lower()
    rows = []
    if fn.endswith('.xls') and not fn.endswith('.xlsx'):
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("xls 파일을 읽으려면 xlrd 라이브러리가 필요합니다 (requirements.txt에 xlrd 추가)")
        data = file_storage.read()
        wb = xlrd.open_workbook(file_contents=data)
        sheet = wb.sheet_by_index(0)
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    else:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active
        for r in ws.values:
            rows.append(list(r))

    header_idx = name_col = amount_col = date_col = -1
    for i, row in enumerate(rows[:20]):
        if not row: continue
        cells = [str(c or '').strip() for c in row]
        ni = next((j for j, c in enumerate(cells) if '보낸' in c or '받는' in c), -1)
        ai = next((j for j, c in enumerate(cells) if c == '입금액(원)' or c == '입금액' or '입금' in c), -1)
        di = next((j for j, c in enumerate(cells) if '거래일' in c or '날짜' in c), -1)
        if ni >= 0 and ai >= 0:
            header_idx = i
            name_col = ni
            amount_col = ai
            date_col = di
            break
    if header_idx < 0:
        return []

    deposits = []
    for row in rows[header_idx + 1:]:
        if not row: continue
        name_v = row[name_col] if name_col < len(row) else ''
        name = str(name_v or '').strip()
        if not name: continue
        try:
            raw = str(row[amount_col] if amount_col < len(row) else '').replace(',', '').replace('원', '').strip()
            if not raw or raw == '0': continue
            amt = int(float(raw))
        except (ValueError, TypeError):
            continue
        if amt <= 0: continue
        date_str = ''
        if date_col >= 0 and date_col < len(row):
            date_str = str(row[date_col] or '').strip()
        deposits.append({'name': name, 'amount': amt, 'date': date_str})
    return deposits


@app.route('/api/bank/upload', methods=['POST'])
def upload_bank():
    """은행 입금내역 엑셀 업로드 → 각 행을 SMS와 동일하게 매칭 (자동확인 or 의심후보)"""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    try:
        deposits = parse_bank_excel(file)
    except Exception as e:
        return jsonify({'error': f'엑셀 파싱 오류: {e}'}), 400

    if not deposits:
        return jsonify({'error': '입금 내역을 찾을 수 없습니다 (헤더에 "보낸분/받는분" + "입금액"이 있어야 함)'}), 400

    matched_high = matched_cand = no_match = 0
    for d in deposits:
        try:
            recv_iso = now_kst().isoformat()
            conn = get_conn()
            cur = conn.execute(
                "INSERT INTO sms_payments (sender, body, parsed_name, parsed_amount, received_at) VALUES (?,?,?,?,?)",
                ('BANK_EXCEL', f"엑셀:{d['name']} {d['amount']:,}원 {d.get('date','')}",
                 d['name'], d['amount'], recv_iso)
            )
            sms_id = cur.lastrowid
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"은행 입금 저장 오류: {e}")
            no_match += 1
            continue
        try:
            parsed = {'name': d['name'], 'amount': d['amount'], 'bank': '엑셀'}
            match_sms_to_order(parsed, recv_iso, sms_id=sms_id)
            conn = get_conn()
            row = conn.execute("SELECT matched FROM sms_payments WHERE id=?", (sms_id,)).fetchone()
            matched = (row['matched'] == 1) if row else False
            if matched:
                matched_high += 1
            else:
                cand = conn.execute(
                    "SELECT 1 FROM match_candidates WHERE source='sms' AND source_ref=? AND status='open'",
                    (str(sms_id),)
                ).fetchone()
                if cand: matched_cand += 1
                else: no_match += 1
            conn.close()
        except Exception as e:
            logger.error(f"은행 입금 매칭 오류: {e}")
            no_match += 1

    logger.info(f"📥 은행입금 업로드: 총 {len(deposits)}건 → 자동확인 {matched_high}, 의심후보 {matched_cand}, 매칭실패 {no_match}")
    return jsonify({
        'ok': True,
        'total': len(deposits),
        'matched_high': matched_high,
        'matched_candidate': matched_cand,
        'no_match': no_match
    })


# ══════════════════════════════════════════════════════════════════
#  닉네임 매핑 CSV Export / Import
# ══════════════════════════════════════════════════════════════════
@app.route('/api/mappings/export', methods=['GET'])
def export_mappings():
    """닉네임 매핑 전체를 텍스트 파일로 다운로드.
    양식:  nickname=realname     (동일인)
           nickname≠realname     (동일인 아님)
    """
    from flask import Response
    conn = get_conn()
    rows = conn.execute(
        'SELECT nickname, realname, COALESCE(negative,0) as negative FROM nick_mappings ORDER BY negative, nickname'
    ).fetchall()
    conn.close()
    lines = []
    for r in rows:
        sep = '≠' if r['negative'] == 1 else '='
        lines.append(f"{r['nickname']}{sep}{r['realname']}")
    text = '\n'.join(lines) + '\n'
    fname = f'nick_mappings_{now_kst().strftime("%Y%m%d_%H%M")}.txt'
    return Response(text,
                    mimetype='text/plain; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@app.route('/api/mappings/import', methods=['POST'])
def import_mappings():
    """텍스트 양식 업로드로 닉네임 매핑 일괄 추가.
    한 줄당 한 매핑:
        nickname=realname     → 동일인 (negative=0)
        nickname≠realname  → 동일인 아님 (negative=1)
        nickname!=realname    → 동일인 아님 (대체 표기)
    빈 줄, # 으로 시작하는 줄은 무시.
    같은 nickname이 이미 있으면 덮어쓰기.
    """
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    try:
        raw = file.read()
        text = None
        for enc in ('utf-8-sig', 'utf-8', 'cp949', 'euc-kr'):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            return jsonify({'error': '파일 인코딩 해석 실패'}), 400
    except Exception as e:
        return jsonify({'error': f'파일 읽기 오류: {e}'}), 400

    added = updated = skipped = 0
    conn = get_conn()
    try:
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith('#'):
                skipped += 1
                continue
            sep = None
            neg = 0
            if '≠' in line:
                sep, neg = '≠', 1
            elif '!=' in line:
                sep, neg = '!=', 1
            elif '=' in line:
                sep, neg = '=', 0
            if not sep:
                skipped += 1
                continue
            parts = line.split(sep, 1)
            if len(parts) != 2:
                skipped += 1
                continue
            nick = parts[0].strip()
            real = parts[1].strip()
            if not nick or not real:
                skipped += 1
                continue
            try:
                existing = conn.execute('SELECT id FROM nick_mappings WHERE nickname=?', (nick,)).fetchone()
                conn.execute(
                    "INSERT OR REPLACE INTO nick_mappings (nickname, realname, negative) VALUES (?,?,?)",
                    (nick, real, neg)
                )
                if existing: updated += 1
                else: added += 1
            except Exception as e:
                skipped += 1
                logger.warning(f"매핑 import 행 스킵: {e}")
        conn.commit()
    finally:
        conn.close()
    logger.info(f"📤 매핑 import (텍스트): 추가 {added}, 수정 {updated}, 스킵 {skipped}")
    return jsonify({'ok': True, 'added': added, 'updated': updated, 'skipped': skipped})


# ══════════════════════════════════════════════════════════════════
#  핵심 매칭 로직 (구매자별 SUM 합산 매칭 + 의심후보 시스템)
# ══════════════════════════════════════════════════════════════════
def _norm(s):
    """이름 정규화: 공백/특수문자 제거, 소문자화"""
    if not s: return ''
    return re.sub(r'[\s\(\)\[\]\.\,\-_]+', '', str(s)).lower()


def resolve_names(conn, name):
    """닉네임 → 실명 매핑 (양방향, negative=1 mapping은 제외)"""
    names = [name]
    if not name: return names
    row = conn.execute(
        "SELECT realname FROM nick_mappings WHERE nickname=? AND COALESCE(negative,0)=0",
        (name,)
    ).fetchone()
    if row and row['realname']: names.append(row['realname'])
    row = conn.execute(
        "SELECT nickname FROM nick_mappings WHERE realname=? AND COALESCE(negative,0)=0",
        (name,)
    ).fetchone()
    if row and row['nickname']: names.append(row['nickname'])
    return list(dict.fromkeys(filter(None, names)))


def amount_matches(paid, ordered):
    diff = abs(paid - ordered)
    return diff == 0 or diff == 4000 or diff <= 100


def find_buyer_match(conn, session_id, search_names, amount):
    """
    퍼지 매칭 - 구매자별 SUM(amount)를 기준으로.
    Returns: ({'buyer_name','total_amount','order_ids'}, confidence, reason)
    """
    if not search_names: return (None, None, '검색이름 없음')
    grouped_rows = conn.execute('''
        SELECT buyer_name, SUM(amount) AS total_amount, COUNT(*) AS row_count,
               GROUP_CONCAT(id) AS order_ids
        FROM orders WHERE session_id=? AND status='pending' GROUP BY buyer_name
    ''', (session_id,)).fetchall()
    if not grouped_rows: return (None, None, '대기 주문 없음')
    grouped = [{
        'buyer_name': r['buyer_name'],
        'total_amount': int(r['total_amount'] or 0),
        'order_ids': [int(x) for x in (r['order_ids'] or '').split(',') if x],
    } for r in grouped_rows]

    # 1) 이름 정확일치 + 합산금액 일치 → high
    for name in search_names:
        for g in grouped:
            if g['buyer_name'] == name and amount_matches(amount, g['total_amount']):
                return (g, 'high', f"이름정확({name})+합산일치")

    # 2) 정규화/부분 + 합산금액 일치 → medium
    norm_searches = {n: _norm(n) for n in search_names if n}
    for name, nname in norm_searches.items():
        if not nname: continue
        for g in grouped:
            buyer = g['buyer_name'] or ''; nbuyer = _norm(buyer)
            if not nbuyer: continue
            if nname == nbuyer and amount_matches(amount, g['total_amount']):
                return (g, 'medium', f"정규화일치({name}↔{buyer})+합산일치")
            if (nname in nbuyer or nbuyer in nname) and amount_matches(amount, g['total_amount']):
                return (g, 'medium', f"부분포함({name}↔{buyer})+합산일치")

    # 3) 합산금액만 일치 (1명) → medium
    amount_only = [g for g in grouped if amount_matches(amount, g['total_amount'])]
    if len(amount_only) == 1:
        only = amount_only[0]
        return (only, 'medium', f"합산금액만 일치({only['buyer_name']})")
    elif len(amount_only) > 1:
        return (amount_only[0], 'low', f"합산금액 같은 {len(amount_only)}명")

    # 4) 이름 유사 + 합산금액 다름 → low
    for name, nname in norm_searches.items():
        if not nname: continue
        for g in grouped:
            buyer = g['buyer_name'] or ''; nbuyer = _norm(buyer)
            if nname == nbuyer or (nname and nbuyer and (nname in nbuyer or nbuyer in nname)):
                return (g, 'low', f"이름유사({name}↔{buyer}) 합산다름({amount:,}↔{g['total_amount']:,})")

    return (None, None, f"후보없음 ({search_names}, {amount:,}원)")


def find_amount_only_candidates(conn, session_id, amount):
    """동일 합산금액 가진 모든 pending 구매자 반환 — 다중 후보 생성용"""
    sql = "SELECT buyer_name, SUM(amount) AS total_amount, GROUP_CONCAT(id) AS order_ids "
    sql += "FROM orders WHERE session_id=? AND status='pending' GROUP BY buyer_name"
    rows = conn.execute(sql, (session_id,)).fetchall()
    out = []
    for r in rows:
        total = int(r['total_amount'] or 0)
        if amount_matches(amount, total):
            ids = [int(x) for x in (r['order_ids'] or '').split(',') if x]
            out.append({'buyer_name': r['buyer_name'], 'total_amount': total, 'order_ids': ids})
    return out


def save_candidate(conn, session_id, source, source_ref,
                   paid_name, paid_name2, amount, paid_at,
                   buyer_dict, confidence, reason):
    if buyer_dict:
        buyer_name = buyer_dict.get('buyer_name')
        cand_amt   = buyer_dict.get('total_amount')
        first_id   = (buyer_dict.get('order_ids') or [None])[0]
    else:
        buyer_name = None; cand_amt = None; first_id = None
    # ⭐ 사용자가 이미 [❌ 거절]한 (paid_name ↔ buyer_name) pair는 다시 후보로 만들지 않음
    if paid_name and buyer_name:
        neg = conn.execute(
            "SELECT 1 FROM nick_mappings WHERE nickname=? AND realname=? AND COALESCE(negative,0)=1",
            (paid_name, buyer_name)
        ).fetchone()
        if neg:
            logger.info(f"  🚫 후보 스킵 (이전 거절): '{paid_name}'↔'{buyer_name}'")
            return
    src_ref_str = str(source_ref) if source_ref else ''
    # v19: 이미 같은 후보가 있으면 UPDATE (이전엔 silent skip → 금액 갱신 안되던 버그)
    existing = conn.execute(
        "SELECT id, status, amount, candidate_amount, confidence FROM match_candidates "
        "WHERE session_id=? AND source=? AND source_ref=? "
        "  AND COALESCE(candidate_order_id,-1)=COALESCE(?,-1)",
        (session_id, source, src_ref_str, first_id)
    ).fetchone()
    if existing:
        if existing['status'] != 'open':
            return  # 사용자가 이미 거절/승인/삭제한 후보는 그대로 둠
        if (existing['amount'] == amount and existing['candidate_amount'] == cand_amt
                and existing['confidence'] == confidence):
            return  # 값 동일하면 스킵
        conn.execute(
            "UPDATE match_candidates SET paid_name=?, paid_name2=?, amount=?, paid_at=?, "
            "candidate_buyer_name=?, candidate_amount=?, confidence=?, reason=?, "
            "created_at=? WHERE id=?",
            (paid_name, paid_name2, amount, paid_at,
             buyer_name, cand_amt, confidence, reason,
             now_kst().isoformat(), existing['id'])
        )
        logger.info(f"  🔄 의심후보 갱신: id={existing['id']} '{paid_name}' {amount:,}원 → '{buyer_name}' 합계{cand_amt} ({confidence})")
        return
    try:
        conn.execute(
            "INSERT INTO match_candidates "
            "(session_id, source, source_ref, paid_name, paid_name2, "
            " amount, paid_at, candidate_order_id, candidate_buyer_name, "
            " candidate_amount, confidence, reason, status, created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,'open',?)",
            (session_id, source, src_ref_str,
             paid_name, paid_name2, amount, paid_at,
             first_id, buyer_name, cand_amt, confidence, reason,
             now_kst().isoformat()))
        logger.info(f"  🤔 의심후보: src={source} '{paid_name}' {amount:,}원 → '{buyer_name}' 합계{cand_amt} ({confidence}: {reason})")
    except Exception as e:
        if 'UNIQUE' not in str(e):
            logger.warning(f"후보 저장 실패: {e}")


def match_sms_to_order(parsed, recv_time, sms_id=None):
    """SMS/은행 입금 한 건 → SUM 합산 기준 매칭.
    [v14] 모든 세션을 끝까지 검사해 어디든 HIGH가 있으면 그걸 자동확인.
          HIGH가 한 곳도 없으면 가장 신뢰도 높은 medium/low를 후보로 저장.
    """
    name   = parsed.get('name')
    amount = parsed['amount']
    now    = now_kst().isoformat()

    conn = get_conn()
    try:
        sessions = conn.execute('''
            SELECT * FROM live_sessions
            WHERE EXISTS (SELECT 1 FROM orders o WHERE o.session_id=live_sessions.id AND o.status='pending')
            ORDER BY live_date DESC
        ''').fetchall()
        if not sessions:
            logger.info(f"SMS: pending 보유 세션 없음 → 보류 ({name or '-'} {amount:,}원)")
            return

        search_names = resolve_names(conn, name) if name else []
        logger.info(f"  🔎 매칭 시작: '{name or '-'}' {amount:,}원 / search_names={search_names} / sessions={[s['id'] for s in sessions]}")

        rank = {'high': 3, 'medium': 2, 'low': 1}
        best_partial = None  # (sid, buyer, confidence, reason)

        # ── Pass 1: 전체 세션 훑어서 HIGH 찾고, 못 찾으면 best partial 트래킹
        for sess in sessions:
            sid = sess['id']
            buyer, confidence, reason = find_buyer_match(
                conn, sid, search_names or [name or ''], amount
            )
            logger.info(f"    session {sid} ({sess['live_date']}): {confidence} | {reason}")
            if confidence == 'high':
                for oid in buyer['order_ids']:
                    conn.execute('''UPDATE orders SET status='confirmed', confirmed_at=?,
                                           pay_type='transfer', bank_date=? WHERE id=?''',
                                 (now, recv_time, oid))
                if sms_id is not None:
                    conn.execute('UPDATE sms_payments SET matched=1 WHERE id=?', (sms_id,))
                else:
                    conn.execute('''UPDATE sms_payments SET matched=1
                                   WHERE parsed_amount=? AND COALESCE(parsed_name,'')=COALESCE(?,'') AND matched=0''',
                                 (amount, name))
                conn.commit()
                logger.info(f"  ✅ SMS 자동매칭: {buyer['buyer_name']} 합계 {buyer['total_amount']:,}원 ({len(buyer['order_ids'])}건) [session={sid}]")
                return
            elif confidence in ('medium', 'low'):
                if best_partial is None or rank.get(confidence, 0) > rank.get(best_partial[2], 0):
                    best_partial = (sid, buyer, confidence, reason)

        # ── Pass 2: HIGH 없음 → 가장 신뢰도 높은 partial을 후보로
        if best_partial is not None:
            sid, buyer, confidence, reason = best_partial
            if confidence == 'low' and '명' in (reason or ''):
                # 동일 합산금액 후보 전부 등록
                extras = find_amount_only_candidates(conn, sid, amount)
                for ex in extras:
                    save_candidate(conn, sid, 'sms', sms_id, name, None, amount, recv_time,
                                   ex, 'low', f"합산금액 동일 ({ex['buyer_name']})")
            else:
                save_candidate(conn, sid, 'sms', sms_id, name, None, amount, recv_time,
                               buyer, confidence, reason)
            conn.commit()
            return

        logger.info(f"  ⚠️ SMS 매칭/후보 모두 실패: '{name or '-'}' {amount:,}원")
    except Exception as e:
        logger.error(f"SMS 매칭 오류: {e}")
    finally:
        conn.close()



def run_auto_check():
    """15분마다 + 수동: 아임웹 카드결제 + 미매칭 SMS 재대조 (SUM 합산 + 의심후보)"""
    logger.info("자동 입금확인 시작...")
    today    = now_kst().strftime('%Y-%m-%d')
    today_ym = today.replace('-', '')
    now_time = now_kst().strftime('%H:%M:%S')

    conn = get_conn()
    try:
        # ⭐ pending 주문이 1건이라도 있는 모든 세션 처리 (check_end 만료 무관)
        sessions = conn.execute('''
            SELECT * FROM live_sessions
            WHERE EXISTS (SELECT 1 FROM orders o WHERE o.session_id=live_sessions.id AND o.status='pending')
            ORDER BY live_date DESC
        ''').fetchall()
        logger.info(f"처리할 세션: {len(sessions)}건 (pending 보유)")

        for session in sessions:
            session = dict(session)
            sid = session['id']
            logger.info(f"세션 확인: {session['filename']}")

            imweb_status = 'success'; sms_status = 'success'
            imweb_confirmed = 0; sms_confirmed = 0; error_msg = None

            # ⭐ v21: stale OPEN imweb 후보 일괄 정리 (옛날 금액 잔존 방지)
            #   사용자가 거절/승인/삭제한 후보(status != 'open')는 보존됨.
            deleted_stale = conn.execute(
                "DELETE FROM match_candidates WHERE source='imweb' AND status='open' AND session_id=?",
                (sid,)
            ).rowcount
            if deleted_stale:
                logger.info(f"  🧹 stale 아임웹 후보 {deleted_stale}건 정리 (재스캔으로 최신 금액으로 재생성됨)")

            # 1) 아임웹 카드결제 — 조회기간 라이브 ~ +30일 (오늘로 캡)
            try:
                live_dt = datetime.strptime(session['live_date'], '%Y-%m-%d')
                start_dt = live_dt
                end_dt = live_dt + timedelta(days=30)
                if end_dt > now_kst(): end_dt = now_kst()
                start_ym = start_dt.strftime('%Y%m%d')
                end_ym   = end_dt.strftime('%Y%m%d')
                imweb_orders = get_paid_orders(start_ym, end_ym)
                logger.info(f"아임웹 조회: {len(imweb_orders)}건 ({start_ym}~{end_ym})")

                for iorder in imweb_orders:
                    info = extract_order_info(iorder)
                    if not info['amount']:
                        continue
                    search_names = []
                    if info.get('name'):  search_names += resolve_names(conn, info['name'])
                    if info.get('name2'): search_names += resolve_names(conn, info['name2'])
                    search_names = list(dict.fromkeys(filter(None, search_names)))

                    order_no = str(iorder.get('order_no') or iorder.get('orderNo') or '')
                    buyer, conf, reason = find_buyer_match(conn, sid, search_names, info['amount'])

                    if conf == 'high':
                        for oid in buyer['order_ids']:
                            conn.execute("UPDATE orders SET status='confirmed', confirmed_at=?, pay_type='card' WHERE id=?",
                                         (info['paid_at'], oid))
                        imweb_confirmed += 1
                        logger.info(f"  ✅ 카드결제: {buyer['buyer_name']} 합계 {buyer['total_amount']:,}원")
                        # ⭐ 아임웹 측 주문도 '상품 준비중' → '배송대기'로 자동 전환
                        try:
                            order_no_v = iorder.get('order_no') or iorder.get('orderNo')
                            prod_list = iorder.get('prod_list') or iorder.get('prod_items') or []
                            prod_order_nos = [p.get('prod_order_no') or p.get('prodOrderNo') for p in prod_list if (p.get('prod_order_no') or p.get('prodOrderNo'))]
                            if order_no_v:
                                set_order_to_standby(str(order_no_v), prod_order_nos or None)
                        except Exception as _e:
                            logger.warning(f"  배송대기 전환 시도 실패: {_e}")
                    elif conf in ('medium', 'low'):
                        save_candidate(conn, sid, 'imweb', order_no,
                                       info.get('name'), info.get('name2'),
                                       info['amount'], info['paid_at'],
                                       buyer, conf, reason)
            except Exception as e:
                imweb_status = 'error'
                error_msg = f"아임웹:{str(e)}"
                logger.error(f"아임웹 조회 오류: {e}")

            # 2) 미매칭 SMS 재대조 — [v14] match_sms_to_order에 위임 (다중세션 high 우선)
            try:
                unmatched_sms = conn.execute("SELECT * FROM sms_payments WHERE matched=0").fetchall()
                conn.commit()  # 위 INSERT 확정 후 새 conn에서도 보이게
                for sms in unmatched_sms:
                    sms = dict(sms)
                    amt = sms.get('parsed_amount') or 0
                    sname = sms.get('parsed_name')
                    if not amt: continue
                    before = conn.execute("SELECT matched FROM sms_payments WHERE id=?", (sms['id'],)).fetchone()
                    match_sms_to_order({'name': sname, 'amount': amt}, sms['received_at'], sms_id=sms['id'])
                    after = conn.execute("SELECT matched FROM sms_payments WHERE id=?", (sms['id'],)).fetchone()
                    if before and after and before[0] == 0 and after[0] == 1:
                        sms_confirmed += 1
            except Exception as e:
                sms_status = 'error'
                error_msg = (error_msg or '') + f" SMS:{str(e)}"
                logger.error(f"SMS 재대조 오류: {e}")

            conn.execute("INSERT INTO check_logs (session_id, check_date, check_time, imweb_status, sms_status, imweb_confirmed, sms_confirmed, error_message) VALUES (?,?,?,?,?,?,?,?)",
                         (sid, today, now_time, imweb_status, sms_status, imweb_confirmed, sms_confirmed, error_msg))

        conn.commit()
        logger.info("자동 입금확인 종료")
    except Exception as e:
        logger.error(f"자동확인 오류: {e}")
        conn.rollback()
    finally:
        conn.close()


# ══════════════════════════════════════════════════════════════════
#  모듈 레벨 초기화 — gunicorn으로 import 되어도 스케줄러가 동작하도록.
# ══════════════════════════════════════════════════════════════════
_SCHEDULER_STARTED = False

def _ensure_scheduler():
    """스케줄러 1회만 시작 (gunicorn 멀티워커 환경에서도 첫 import 때 한 번)."""
    global _SCHEDULER_STARTED
    if _SCHEDULER_STARTED:
        return
    try:
        sched = BackgroundScheduler(timezone='Asia/Seoul')
        # 첫 실행: 서버 기동 10초 뒤 즉시 1회 → 이후 15분마다
        sched.add_job(run_auto_check, 'interval', minutes=15,
                      id='interval_check', replace_existing=True,
                      next_run_time=now_kst() + timedelta(seconds=10))
        sched.start()
        _SCHEDULER_STARTED = True
        logger.info("⏰ 스케줄러 시작 (10초 후 1회 + 15분마다)")
    except Exception as e:
        logger.error(f"스케줄러 시작 실패: {e}")

# ── v25 변경: 세션삭제 API / 위탁정산 구매자별·위탁전체인식 / 정산엑셀 폼(제목·날짜·맑은고딕) ──
# DB 초기화 + 스케줄러 시작 (모듈 import 시점)
init_db()
_ensure_scheduler()

if __name__ == '__main__':
    PORT = int(os.environ.get('PORT', 5000))
    logger.info(f"🌿 지양하월시아 서버 시작! 포트: {PORT}")
    app.run(host='0.0.0.0', port=PORT, debug=False, use_reloader=False)
